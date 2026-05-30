import os
import shutil
import smtplib
from datetime import datetime
from email.mime.text import MIMEText
from io import BytesIO

from flask import current_app
from openpyxl import Workbook

from .models import Customer, Device, Line, RemoteAccess, Renewal, User, db


def parse_date(value):
    if not value:
        return None
    return datetime.strptime(value, "%Y-%m-%d").date()


def money(value):
    try:
        return round(float(value or 0), 2)
    except ValueError:
        return 0


def create_backup():
    uri = current_app.config["SQLALCHEMY_DATABASE_URI"]
    if not uri.startswith("sqlite:///"):
        raise RuntimeError("当前备份脚本仅支持 SQLite")
    db_path = uri.replace("sqlite:///", "", 1)
    if not os.path.isabs(db_path):
        db_path = os.path.abspath(db_path)
    backup_dir = os.path.abspath(os.path.join(current_app.root_path, "..", "backups"))
    os.makedirs(backup_dir, exist_ok=True)
    target = os.path.join(backup_dir, f"app-{datetime.now().strftime('%Y%m%d-%H%M%S')}.db")
    db.session.remove()
    shutil.copy2(db_path, target)
    cleanup_backups(backup_dir)
    return target


def cleanup_backups(backup_dir, keep=30):
    files = sorted(
        [os.path.join(backup_dir, name) for name in os.listdir(backup_dir) if name.endswith(".db")],
        key=os.path.getmtime,
        reverse=True,
    )
    for path in files[keep:]:
        os.remove(path)


def list_backups():
    backup_dir = os.path.abspath(os.path.join(current_app.root_path, "..", "backups"))
    os.makedirs(backup_dir, exist_ok=True)
    files = []
    for name in os.listdir(backup_dir):
        if name.endswith(".db"):
            path = os.path.join(backup_dir, name)
            files.append({"name": name, "size": os.path.getsize(path), "mtime": datetime.fromtimestamp(os.path.getmtime(path))})
    return sorted(files, key=lambda item: item["mtime"], reverse=True)


def export_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "客户"
    ws.append(["客户名称", "联系人", "微信", "Telegram", "邮箱", "手机号", "状态", "备注", "创建时间", "更新时间"])
    for item in Customer.query.order_by(Customer.id.desc()).all():
        ws.append([item.name, item.contact, item.wechat, item.telegram, item.email, item.phone, item.status, item.remark, item.created_at, item.updated_at])

    ws = wb.create_sheet("设备")
    ws.append(["设备名称", "所属客户", "国家", "设备类型", "备注", "创建时间", "更新时间"])
    for item in Device.query.order_by(Device.id.desc()).all():
        ws.append([item.name, item.customer.name if item.customer else "", item.country, item.device_type, item.remark, item.created_at, item.updated_at])

    ws = wb.create_sheet("线路")
    ws.append(["线路编号", "国家", "线路类型", "备注", "创建时间", "更新时间"])
    for item in Line.query.order_by(Line.id.desc()).all():
        ws.append([item.code, item.country, item.line_type, item.remark, item.created_at, item.updated_at])

    ws = wb.create_sheet("续费记录")
    ws.append(["客户", "设备", "线路", "价格", "开通时间", "到期时间", "状态", "备注"])
    for item in Renewal.query.order_by(Renewal.expire_date.asc()).all():
        ws.append([
            item.customer.name if item.customer else "",
            item.device.name if item.device else "",
            item.line.code if item.line else "",
            float(item.price or 0),
            item.start_date,
            item.expire_date,
            item.computed_status,
            item.remark,
        ])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def send_reminder_email():
    notices = [r for r in Renewal.query.all() if r.status != "已续费" and (r.days_left in [7, 3, 1] or r.days_left < 0)]
    if not notices:
        return 0
    admin = User.query.first()
    to_addr = os.getenv("ADMIN_EMAIL") or (admin.email if admin else "")
    if not to_addr:
        return 0

    lines = []
    for item in notices:
        lines.append(
            "\n".join(
                [
                    f"客户名称：{item.customer.name if item.customer else ''}",
                    f"设备名称：{item.device.name if item.device else ''}",
                    f"线路名称：{item.line.code if item.line else ''}",
                    f"到期时间：{item.expire_date}",
                    f"剩余天数：{item.days_left}",
                    f"价格：{item.price}",
                    f"备注：{item.remark or ''}",
                ]
            )
        )
    body = "\n\n---\n\n".join(lines)
    msg = MIMEText(body, "plain", "utf-8")
    msg["Subject"] = "【续费提醒】"
    msg["From"] = os.getenv("MAIL_FROM", os.getenv("MAIL_USERNAME", ""))
    msg["To"] = to_addr

    server = os.getenv("MAIL_SERVER")
    username = os.getenv("MAIL_USERNAME")
    password = os.getenv("MAIL_PASSWORD")
    if not server or not username or not password:
        return 0
    port = int(os.getenv("MAIL_PORT", "587"))
    with smtplib.SMTP(server, port, timeout=20) as smtp:
        if os.getenv("MAIL_USE_TLS", "true").lower() == "true":
            smtp.starttls()
        smtp.login(username, password)
        smtp.send_message(msg)
    return len(notices)


def global_search(query):
    like = f"%{query}%"
    return {
        "customers": Customer.query.filter((Customer.name.like(like)) | (Customer.remark.like(like))).all(),
        "devices": Device.query.filter((Device.name.like(like)) | (Device.remark.like(like))).all(),
        "lines": Line.query.filter((Line.code.like(like)) | (Line.remark.like(like))).all(),
        "remotes": RemoteAccess.query.filter((RemoteAccess.address.like(like)) | (RemoteAccess.remark.like(like))).all(),
    }
